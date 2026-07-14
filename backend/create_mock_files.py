# create_mock_files.py
import openpyxl
import os

def create_mocks():
    print("Creating mock Excel data files...")
    
    # 1. Create Student PIN List Excel file
    wb_students = openpyxl.Workbook()
    ws_cn = wb_students.active
    ws_cn.title = "Computer Networks"
    ws_cn.append(["S.No", "Roll Number", "Subject Name"])
    # 35 Students for CSE Computer Networks
    for i in range(1, 36):
        ws_cn.append([i, f"23711A05{str(i).zfill(2)}", "Computer Networks"])
        
    ws_se = wb_students.create_sheet(title="Software Engineering")
    ws_se.append(["S.No", "Roll Number", "Subject Name"])
    # 28 Students for CST Software Engineering
    for i in range(1, 29):
        ws_se.append([i, f"23711A12{str(i).zfill(2)}", "Software Engineering"])

    ws_dsp = wb_students.create_sheet(title="Digital Signal Processing")
    ws_dsp.append(["S.No", "Roll Number", "Subject Name"])
    # 10 Students for ECE Digital Signal Processing
    for i in range(1, 11):
        ws_dsp.append([i, f"23711A04{str(i).zfill(2)}", "Digital Signal Processing"])
        
    students_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mock_students_list.xlsx")
    wb_students.save(students_path)
    print(f"Created student list at: {students_path}")
    
    # 2. Create Master Exam Schedule Excel file
    wb_schedule = openpyxl.Workbook()
    ws_sched = wb_schedule.active
    ws_sched.title = "Exam Schedule"
    ws_sched.append(["Exam Date", "Exam Time", "Subject Name"])
    
    # Add slot 1: July 15, 2026, 10:00 AM - 01:00 PM (CN and SE scheduled)
    ws_sched.append(["2026-07-15", "10:00 AM - 01:00 PM", "Computer Networks"])
    ws_sched.append(["2026-07-15", "10:00 AM - 01:00 PM", "Software Engineering"])
    
    # Add slot 2: July 16, 2026, 01:30 PM - 04:30 PM (DSP scheduled)
    ws_sched.append(["2026-07-16", "01:30 PM - 04:30 PM", "Digital Signal Processing"])
    
    schedule_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mock_exam_schedule.xlsx")
    wb_schedule.save(schedule_path)
    print(f"Created exam schedule at: {schedule_path}")

if __name__ == "__main__":
    create_mocks()
