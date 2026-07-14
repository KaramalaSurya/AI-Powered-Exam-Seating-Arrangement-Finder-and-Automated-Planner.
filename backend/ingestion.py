# ingestion.py
import re
import io
import csv
import openpyxl
import pdfplumber
from typing import List, Dict, Any, Tuple

def normalize_date_string(date_str: str) -> str:
    """
    Normalizes day-month-year or year-month-day string formats into YYYY-MM-DD.
    Supports dots (.), slashes (/), and dashes (-) as delimiters.
    """
    # Remove time part if present (e.g. 2026-07-15 00:00:00)
    date_str = str(date_str).strip().split(' ')[0]
    # Replace dots and slashes with dashes
    normalized = re.sub(r'[./]', '-', date_str)
    
    # Check if format is DD-MM-YYYY (or single digits D-M-YYYY)
    match_dmw = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', normalized)
    if match_dmw:
        day, month, year = match_dmw.groups()
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
    return normalized


def parse_student_pin_list(file_bytes: bytes, filename: str) -> List[Dict[str, str]]:
    """
    Parses an uploaded Student Registration / PIN List (PDF or Excel).
    Returns a list of dicts: [{"roll_number": "...", "subject": "..."}]
    """
    ext = filename.split('.')[-1].lower()
    students = []

    # Roll number regex (MITS standard is 10 alphanumeric characters e.g. 23711A0518, 24691A05R0)
    # Usually starts with 2 digits, then 5 characters (e.g. 711A0, 691A0), then 3 characters.
    roll_pattern = re.compile(r'\b\d{2}[A-Z0-9]{8}\b', re.IGNORECASE)

    if ext in ['xlsx', 'xls']:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            for sheet in wb.worksheets:
                sheet_name = sheet.title.strip()
                
                # Check headers
                headers = []
                roll_col_idx = -1
                subj_col_idx = -1
                
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if not any(row):
                        continue
                    
                    # Convert cells to string
                    row_strs = [str(c).strip() if c is not None else "" for c in row]
                    
                    # Look for headers in first few rows
                    if r_idx < 5 and (roll_col_idx == -1 or subj_col_idx == -1):
                        for c_idx, val in enumerate(row_strs):
                            val_lower = val.lower()
                            if any(h in val_lower for h in ["roll", "pin", "student id", "htno", "hall ticket"]):
                                roll_col_idx = c_idx
                            if any(h in val_lower for h in ["subject", "course", "paper", "branch"]):
                                subj_col_idx = c_idx
                        if roll_col_idx != -1:
                            headers = row_strs
                            continue # Skip header row from student parsing
                    
                    # Parse student data
                    if roll_col_idx != -1 and roll_col_idx < len(row_strs):
                        roll = row_strs[roll_col_idx]
                        if roll_pattern.match(roll):
                            subject = "General"
                            if subj_col_idx != -1 and subj_col_idx < len(row_strs) and row_strs[subj_col_idx]:
                                subject = row_strs[subj_col_idx]
                            elif sheet_name and not sheet_name.lower().startswith("sheet"):
                                subject = sheet_name
                            
                            students.append({
                                "roll_number": roll.upper(),
                                "subject": subject
                            })
                    else:
                        # Fallback heuristic: search cells for roll number
                        for c_idx, cell_val in enumerate(row_strs):
                            if roll_pattern.match(cell_val):
                                # Found a roll, look for subject in subsequent cells or sheet name
                                roll = cell_val
                                subject = sheet_name if not sheet_name.lower().startswith("sheet") else "General"
                                # Look for a text cell that doesn't look like serial/roll as subject
                                for other_val in row_strs:
                                    if other_val and other_val != roll and not roll_pattern.match(other_val) and len(other_val) > 4:
                                        if not other_val.isdigit():
                                            subject = other_val
                                            break
                                students.append({
                                    "roll_number": roll.upper(),
                                    "subject": subject
                                })
                                break
        except Exception as e:
            print(f"Error parsing Excel: {e}")
            
    elif ext == 'pdf':
        try:
            current_subject = "General"
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    
                    lines = text.split('\n')
                    for line in lines:
                        line = line.strip()
                        if not line:
                            continue
                        
                        subj_match = re.search(r'(?:subject|course|paper|branch)\s*:\s*([^\n\r|]+)', line, re.IGNORECASE)
                        if subj_match:
                            current_subject = subj_match.group(1).strip()
                            continue
                            
                        # If a line has a roll number, extract it
                        rolls_on_line = roll_pattern.findall(line)
                        if rolls_on_line:
                            for roll in rolls_on_line:
                                students.append({
                                    "roll_number": roll.upper(),
                                    "subject": current_subject
                                })
        except Exception as e:
            print(f"Error parsing PDF: {e}")
            
    return students


def parse_master_schedule(file_bytes: bytes, filename: str) -> List[Dict[str, str]]:
    """
    Parses an uploaded Master Exam Schedule (Excel, CSV, or PDF).
    Returns a list of dicts: [{"exam_date": "...", "exam_time": "...", "subject": "..."}]
    """
    ext = filename.split('.')[-1].lower()
    schedule = []

    if ext in ['xlsx', 'xls']:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            for sheet in wb.worksheets:
                date_col_idx = -1
                time_col_idx = -1
                subj_col_idx = -1
                
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if not any(row):
                        continue
                    row_strs = [str(c).strip() if c is not None else "" for c in row]
                    
                    # Detect headers in first few rows
                    if r_idx < 5 and (date_col_idx == -1 or time_col_idx == -1 or subj_col_idx == -1):
                        for c_idx, val in enumerate(row_strs):
                            val_lower = val.lower()
                            if any(h in val_lower for h in ["date"]):
                                date_col_idx = c_idx
                            if any(h in val_lower for h in ["time", "session", "slot"]):
                                time_col_idx = c_idx
                            if any(h in val_lower for h in ["subject", "course", "paper", "exam"]):
                                subj_col_idx = c_idx
                        if date_col_idx != -1 or time_col_idx != -1 or subj_col_idx != -1:
                            continue # Skip header row
                    
                    if date_col_idx != -1 and subj_col_idx != -1:
                        date_val = row_strs[date_col_idx] if date_col_idx < len(row_strs) else ""
                        time_val = row_strs[time_col_idx] if (time_col_idx != -1 and time_col_idx < len(row_strs)) else "10:00 AM - 01:00 PM"
                        subj_val = row_strs[subj_col_idx] if subj_col_idx < len(row_strs) else ""
                        
                        if date_val and subj_val and not any(h in date_val.lower() for h in ["date", "header"]):
                            schedule.append({
                                "exam_date": normalize_date_string(date_val),
                                "exam_time": time_val if time_val else "10:00 AM - 01:00 PM",
                                "subject": subj_val
                            })
        except Exception as e:
            print(f"Error parsing Excel Schedule: {e}")
            
    elif ext in ['csv', 'txt']:
        try:
            content = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
            
            date_col_idx = -1
            time_col_idx = -1
            subj_col_idx = -1
            
            for r_idx, row in enumerate(rows):
                if not any(row):
                    continue
                row_strs = [str(c).strip() for c in row]
                
                # Detect headers
                if r_idx < 5 and (date_col_idx == -1 or time_col_idx == -1 or subj_col_idx == -1):
                    for c_idx, val in enumerate(row_strs):
                        val_lower = val.lower()
                        if any(h in val_lower for h in ["date"]):
                            date_col_idx = c_idx
                        if any(h in val_lower for h in ["time", "session", "slot"]):
                            time_col_idx = c_idx
                        if any(h in val_lower for h in ["subject", "course", "paper", "exam"]):
                            subj_col_idx = c_idx
                    if date_col_idx != -1 or time_col_idx != -1 or subj_col_idx != -1:
                        continue # Skip header row
                
                if date_col_idx != -1 and subj_col_idx != -1:
                    date_val = row_strs[date_col_idx] if date_col_idx < len(row_strs) else ""
                    time_val = row_strs[time_col_idx] if (time_col_idx != -1 and time_col_idx < len(row_strs)) else "10:00 AM - 01:00 PM"
                    subj_val = row_strs[subj_col_idx] if subj_col_idx < len(row_strs) else ""
                    
                    if date_val and subj_val and not any(h in date_val.lower() for h in ["date", "header"]):
                        schedule.append({
                            "exam_date": normalize_date_string(date_val),
                            "exam_time": time_val if time_val else "10:00 AM - 01:00 PM",
                            "subject": subj_val
                        })
        except Exception as e:
            print(f"Error parsing CSV Schedule: {e}")
            
    elif ext == 'pdf':
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    
                    lines = text.split('\n')
                    for line in lines:
                        line = line.strip()
                        if not line:
                            continue
                        
                        # Match Date: YYYY-MM-DD or DD-MM-YYYY or DD.MM.YYYY or similar
                        date_match = re.search(r'\b(\d{4}[-./]\d{2}[-./]\d{2})|(\d{1,2}[-./]\d{1,2}[-./]\d{4})\b', line)
                        if not date_match:
                            continue
                            
                        exam_date_raw = date_match.group(0)
                        exam_date = normalize_date_string(exam_date_raw)
                        
                        # Match Time Slot: e.g. 10:00 AM - 01:00 PM or 01:30 PM TO 03:00 PM
                        time_match = re.search(r'\b(\d{1,2}:\d{2}\s*(?:AM|PM)?\s*(?:-|to|TO)\s*\d{1,2}:\d{2}\s*(?:AM|PM)?)\b', line, re.IGNORECASE)
                        if time_match:
                            exam_time = time_match.group(0).strip()
                        else:
                            exam_time = "10:00 AM - 01:00 PM"
                            
                        # Subject is the remainder of the line
                        subject_part = line
                        subject_part = subject_part.replace(exam_date_raw, "")
                        if time_match:
                            subject_part = subject_part.replace(time_match.group(0), "")
                            
                        # Clean up subject string
                        subject_part = re.sub(r'\b(?:AM|PM|AN|FN)\b', '', subject_part, flags=re.IGNORECASE)
                        subject_part = re.sub(r'^[\s\d\-\|,\.\:\*\/]+', '', subject_part)
                        subject_part = re.sub(r'[\s\-\|,\.\:\*\/]+$', '', subject_part)
                        subject_part = re.sub(r'\s+', ' ', subject_part).strip()
                        
                        if subject_part and len(subject_part) > 3 and not subject_part.lower() in ["subject", "subject name", "course"]:
                            schedule.append({
                                "exam_date": exam_date,
                                "exam_time": exam_time,
                                "subject": subject_part
                            })
        except Exception as e:
            print(f"Error parsing PDF Schedule: {e}")
            
    return schedule

